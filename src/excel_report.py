"""
excel_report.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

class ExcelReport:
    def __init__(self):
        self.workbook=Workbook()
        self.summary=self.workbook.active
        self.summary.title="Summary"
        self.summary.append(["Fund","Bull Markets","Bear Markets","Average Bull %","Average Bear %","Longest Bull (days)","Longest Bear (days)"])
        for c in self.summary[1]:
            c.font=Font(bold=True)

    def add_fund(self,symbol,cycles):
        ws=self.workbook.create_sheet(symbol)
        ws.append(["#","Type","Start","End","Days","Return %"])
        for c in ws[1]:
            c.font=Font(bold=True)
        bulls=[]; bears=[]; lb=0; lr=0
        for i,cycle in enumerate(cycles,1):
            ws.append([i,cycle.cycle_type,str(cycle.start_date.date()),str(cycle.end_date.date()),cycle.days,round(cycle.percent_change,2)])
            if cycle.cycle_type=="Bull":
                bulls.append(cycle.percent_change); lb=max(lb,cycle.days)
            else:
                bears.append(cycle.percent_change); lr=max(lr,cycle.days)
        for col in ws.columns:
            w=max(len(str(x.value)) if x.value is not None else 0 for x in col)+3
            ws.column_dimensions[col[0].column_letter].width=w
        ws.freeze_panes="A2"
        self.summary.append([symbol,len(bulls),len(bears),round(sum(bulls)/len(bulls),2) if bulls else 0,round(sum(bears)/len(bears),2) if bears else 0,lb,lr])

    def save(self,filename="MarketCycleReport.xlsx"):
        for col in self.summary.columns:
            w=max(len(str(x.value)) if x.value is not None else 0 for x in col)+3
            self.summary.column_dimensions[col[0].column_letter].width=w
        self.summary.freeze_panes="A2"
        root=Path(__file__).resolve().parents[3]
        outdir=root/"reports"
        outdir.mkdir(exist_ok=True)
        outfile=outdir/filename
        self.workbook.save(outfile)
        return outfile
